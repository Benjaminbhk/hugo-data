import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook

def main():
    # Initialisation de Tkinter
    root = tk.Tk()
    root.withdraw()

    # Sélection des fichiers Excel à traiter
    file_paths = filedialog.askopenfilenames(
        title="Sélectionnez les fichiers Excel",
        filetypes=[("Fichiers Excel", "*.xlsx")]
    )
    if not file_paths:
        print("Aucun fichier sélectionné. Fin du programme.")
        return

    # Chargement et concaténation des fichiers Excel
    dataframes = []
    for file in file_paths:
        try:
            df = pd.read_excel(file, engine="openpyxl")
            dataframes.append(df)
        except Exception as e:
            print(f"Erreur lors du chargement de {file}: {e}")
    if not dataframes:
        print("Aucun fichier valide n'a été chargé.")
        return
    final_df = pd.concat(dataframes, ignore_index=True)

    # Tri initial par heure si la colonne 'Time' existe
    if 'Time' in final_df.columns:
        final_df['Time'] = pd.to_datetime(final_df['Time'], format='%H:%M:%S', errors='coerce').dt.time
        final_df['sort_order'] = final_df['Time'].apply(lambda x: 0 if x >= pd.Timestamp("08:00:00").time() else 1)
        final_df = final_df.sort_values(by=['sort_order', 'Time']).drop(columns=['sort_order']).reset_index(drop=True)
        # Inversion de l'ordre des lignes : le bas devient le haut, et vice-versa
        final_df = final_df.iloc[::-1].reset_index(drop=True)
    else:
        print("La colonne 'Time' est introuvable dans les fichiers.")
        return

    # Saisie de la date
    date_input = input("Entrez la date au format YYYY-MM-DD : ")
    try:
        trade_date = pd.to_datetime(date_input).date()
    except Exception as e:
        print("Format de date invalide. Fin du programme.")
        return

    # Initialisation des colonnes
    final_df['Date'] = trade_date
    final_df['Structure_ID'] = ""
    final_df['Price'] = pd.to_numeric(final_df['Price'], errors='coerce')
    final_df['Size'] = pd.to_numeric(final_df['Size'], errors='coerce')

    # Séparation des lignes selon la présence de Price
    df_price_na = final_df[final_df['Price'].isna()].copy()
    df_price_ok = final_df[final_df['Price'].notna()].copy()
    df_price_ok['DateTime'] = df_price_ok.apply(lambda row: pd.Timestamp.combine(trade_date, row['Time']), axis=1)

##
    # Détection combinée des paires de Roll (priorité à 120 sec, sinon seuil étendu 10000 sec)
    roll_counter = 0
    for i in range(len(df_price_ok)):
        if df_price_ok.iloc[i]['Structure_ID'] != "":
            continue
        row_i = df_price_ok.iloc[i]

        # Vérification : si le ticker a exactement 7 caractères, c'est déjà un roll groupé
        if len(str(row_i['Ticker'])) == 7:
            roll_counter += 1
            roll_code = f"{trade_date.strftime('%Y%m%d')}-R-{roll_counter}-L0"
            idx_i = df_price_ok.index[i]
            final_df.loc[idx_i, 'Structure_ID'] = roll_code
            df_price_ok.loc[idx_i, 'Structure_ID'] = roll_code
            continue  # Passe à la prochaine itération

        ticker_prefix_i = str(row_i['Ticker'])[:3]
        size_i = row_i['Size']
        price_i = row_i['Price']
        time_i = row_i['DateTime']

        # Variables pour stocker un candidat avec seuil 120 et un candidat étendu
        candidate_j_120 = None
        candidate_j_extended = None

        for j in range(i+1, len(df_price_ok)):
            if df_price_ok.iloc[j]['Structure_ID'] != "":
                continue
            row_j = df_price_ok.iloc[j]
            if str(row_j['Ticker'])[:3] != ticker_prefix_i:
                continue
            if str(row_j['Ticker']) == str(row_i['Ticker']):
                continue
            if abs(size_i - row_j['Size']) > 0.05 * size_i:
                continue
            if abs(price_i - row_j['Price']) > 0.05 * price_i:
                continue

            time_diff = abs((row_j['DateTime'] - time_i).total_seconds())
            if time_diff <= 120:
                candidate_j_120 = j
                break  # On prend immédiatement le candidat qui satisfait le seuil 120
            elif time_diff <= 10000 and candidate_j_extended is None:
                candidate_j_extended = j  # On enregistre le premier candidat sous seuil étendu

        # Choix du candidat : priorité au seuil 120
        if candidate_j_120 is not None:
            chosen_j = candidate_j_120
        elif candidate_j_extended is not None:
            chosen_j = candidate_j_extended
        else:
            continue  # Pas de candidat trouvé pour cette ligne

        roll_counter += 1
        roll_code_leg1 = f"{trade_date.strftime('%Y%m%d')}-R-{roll_counter}-L1"
        roll_code_leg2 = f"{trade_date.strftime('%Y%m%d')}-R-{roll_counter}-L2"
        idx_i = df_price_ok.index[i]
        idx_j = df_price_ok.index[chosen_j]
        final_df.loc[idx_i, 'Structure_ID'] = roll_code_leg1
        final_df.loc[idx_j, 'Structure_ID'] = roll_code_leg2
        df_price_ok.loc[idx_i, 'Structure_ID'] = roll_code_leg1
        df_price_ok.loc[idx_j, 'Structure_ID'] = roll_code_leg2


    # Attribution des labels "Screen" et "Outright"
    final_df.loc[final_df['Price'].isna(), 'Structure_ID'] = f"{trade_date.strftime('%Y%m%d')}-S"
    final_df.loc[final_df['Structure_ID'] == "", 'Structure_ID'] = f"{trade_date.strftime('%Y%m%d')}-O"

    # Création de la colonne "Structure"
    def extract_Structure(struct_code):
        if 'R' in struct_code:
            return 'Leg'
        elif 'S' in struct_code:
            return 'Screen'
        elif 'O' in struct_code:
            return 'Outright'
        else:
            return 'Autre'
    final_df['Structure'] = final_df['Structure_ID'].apply(extract_Structure)

    # Regroupement par catégorie et tri global
    order_mapping = {'Leg': 0, 'Screen': 1, 'Outright': 2, 'Autre': 3}
    final_df['sort_order'] = final_df['Structure'].map(order_mapping)
    final_df = final_df.sort_values(by='sort_order').drop(columns=['sort_order'])
    roll_mask = final_df['Structure_ID'].str.contains("-R-")
    screen_mask = final_df['Structure_ID'].str.contains("-S")
    outright_mask = final_df['Structure_ID'].str.contains("-O")
    df_roll = final_df[roll_mask].copy()
    df_screen = final_df[screen_mask].copy()
    df_outright = final_df[outright_mask].copy()

    # Tri personnalisé des Roll selon le ticker
    df_roll['roll_counter'] = df_roll['Structure_ID'].str.extract(r'-R-(\d+)-L').astype(int)
    df_roll['ticker_last_digit'] = df_roll['Ticker'].str[-1].astype(int)
    df_roll['ticker_penult'] = df_roll['Ticker'].str[-2]
    order_map_letters = {'H': 1, 'M': 2, 'U': 3, 'Z': 4}
    df_roll['ticker_penult_order'] = df_roll['ticker_penult'].map(order_map_letters).fillna(99)
    df_roll_sorted = df_roll.sort_values(by=['roll_counter', 'ticker_last_digit', 'ticker_penult_order'])
    df_screen_sorted = df_screen.sort_values(by='Time')
    df_outright_sorted = df_outright.sort_values(by='Time')
    final_sorted = pd.concat([df_roll_sorted, df_screen_sorted, df_outright_sorted], ignore_index=True)
    final_sorted = final_sorted.drop(columns=['roll_counter', 'ticker_last_digit', 'ticker_penult', 'ticker_penult_order'], errors='ignore')

    # Insertion des lignes résumé pour les Roll (Merge Roll) avec calcul "Level"
    summary_rows = []
    for r in df_roll_sorted['roll_counter'].unique():
        group = df_roll_sorted[df_roll_sorted['roll_counter'] == r]
        if len(group) >= 2:
            row1 = group.iloc[0]
            row2 = group.iloc[1]
            summary = {
                'Time': row1['Time'],
                'Level': (row2['Price'] / row1['Price']-1) * 100,
                'Ticker': row1['Ticker'] + row2['Ticker'][-2:],
                'Notional': (row1['Notional'] + row2['Notional']) / 2,
                'Size': row1['Size'],
                'Price': row1['Price'],
                'Volume': row1['Volume'],
                '1DChg': row1['1DChg'],
                'UndTkr': row1['UndTkr'],
                '1PtVal': row1['1PtVal'],
                'Exch': row1['Exch'],
                'UndCmpName': row1['UndCmpName'],
                'UndPrc': row1['UndPrc'],
                'Date': row1['Date'],
                'FutName': row1['FutName'] + row2['FutName'][-5:],
                'Structure_ID': f"{trade_date.strftime('%Y%m%d')}-R-{r}",
                'Structure': "Roll",
                'roll_counter': r,
                'order': 0
            }
            summary_rows.append(summary)
    df_summary = pd.DataFrame(summary_rows)
    legs_rows = []
    for _, row in df_roll_sorted.iterrows():
        row_copy = row.copy()
        row_copy['order'] = 1 if row_copy['Structure_ID'].endswith("-L1") else 2
        legs_rows.append(row_copy)
    df_legs = pd.DataFrame(legs_rows)
    df_roll_final = pd.concat([df_summary, df_legs], ignore_index=True)
    df_roll_final = df_roll_final.sort_values(by=['roll_counter', 'order'])
    df_roll_final = df_roll_final.drop(columns=['roll_counter', 'order', 'ticker_last_digit', 'ticker_penult', 'ticker_penult_order'], errors='ignore')
    # Traitement des lignes dont la Structure_ID contient "-L0"
    mask_l0 = df_roll_final['Structure_ID'].str.contains("-L0", na=False)
    df_roll_final.loc[mask_l0, 'Structure_ID'] = df_roll_final.loc[mask_l0, 'Structure_ID'].str.replace("-L0", "", regex=False)
    df_roll_final.loc[mask_l0, 'Structure'] = "Roll"

    # Conversion du format de la colonne Date en "MM/DD/YYYY"
    final_df['Date'] = pd.to_datetime(final_df['Date']).dt.strftime('%m/%d/%Y')



    # Assemblage final et export
    final_sorted = pd.concat([df_roll_final, df_screen_sorted, df_outright_sorted], ignore_index=True)

    # Insertion de la colonne "Closing1d" juste après "Price"
    price_idx = final_sorted.columns.get_loc("Price")
    final_sorted.insert(price_idx+1, "Closing1d", "")

    # Remplissage de la colonne "Closing1d" avec la formule Excel pour les lignes "Roll" ou "Outright"
    for i, row in final_sorted.iterrows():
        excel_row = i + 2  # La première ligne de données dans Excel est la ligne 2 (ligne 1 = en-tête)
        if row["Structure"] in ["Roll", "Outright"]:
            final_sorted.at[i, "Closing1d"] = f'=BDH(J{excel_row}&" Index", "PX_CLOSE_1D",O{excel_row},O{excel_row})'
        else:
            final_sorted.at[i, "Closing1d"] = ""

    save_path = filedialog.asksaveasfilename(
        title="Enregistrez le fichier final",
        defaultextension=".xlsx",
        filetypes=[("Fichiers Excel", "*.xlsx")]
    )
    if save_path:
        with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
            final_sorted.to_excel(writer, index=False, sheet_name='Sheet1')
        # Post-traitement avec openpyxl pour réécrire les formules de la colonne "Closing1d"
        wb = load_workbook(save_path)
        ws = wb.active
        closing1d_col_idx = final_sorted.columns.get_loc("Closing1d") + 1  # openpyxl utilise un index 1-based
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=closing1d_col_idx)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                # Réécrire la formule exactement pour éviter l'ajout de "@" par Excel
                ws.cell(row=row_num, column=closing1d_col_idx).value = f'=BDH(J{row_num}&" Index", "PX_CLOSE_1D",O{row_num},O{row_num})'
        level_col_idx = final_sorted.columns.get_loc("Level") + 1
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=level_col_idx)
            cell.number_format = "0.00%"
        wb.save(save_path)
        print(f"Fichier enregistré sous : {save_path}")
    else:
        print("Aucun emplacement de sauvegarde sélectionné.")

if __name__ == "__main__":
    main()
