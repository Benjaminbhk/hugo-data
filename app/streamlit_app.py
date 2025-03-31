import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import io
import os
import calendar
from datetime import date, datetime

# ------------------------
# Fonctions de traitement
# ------------------------

def process_files(uploaded_files, trade_date):
    dataframes = []
    for uploaded_file in uploaded_files:
        try:
            df = pd.read_excel(uploaded_file, engine="openpyxl")
            dataframes.append(df)
        except Exception as e:
            st.error(f"Erreur lors du chargement de {uploaded_file.name}: {e}")
    if not dataframes:
        st.error("Aucun fichier valide n'a été chargé.")
        return None

    # Concaténation de tous les fichiers
    final_df = pd.concat(dataframes, ignore_index=True)

    # Tri initial par heure si la colonne 'Time' existe
    if 'Time' in final_df.columns:
        final_df['Time'] = pd.to_datetime(final_df['Time'], format='%H:%M:%S', errors='coerce').dt.time
        final_df['sort_order'] = final_df['Time'].apply(lambda x: 0 if x >= pd.Timestamp("08:00:00").time() else 1)
        final_df = final_df.sort_values(by=['sort_order', 'Time']).drop(columns=['sort_order']).reset_index(drop=True)
        # Inversion de l'ordre des lignes
        final_df = final_df.iloc[::-1].reset_index(drop=True)
    else:
        st.error("La colonne 'Time' est introuvable dans les fichiers.")
        return None

    # Saisie de la date et initialisation des colonnes
    final_df['Date'] = trade_date
    final_df['Structure_ID'] = ""
    final_df['Price'] = pd.to_numeric(final_df['Price'], errors='coerce')
    final_df['Size'] = pd.to_numeric(final_df['Size'], errors='coerce')

    # Séparation des lignes selon la présence de Price
    df_price_na = final_df[final_df['Price'].isna()].copy()
    df_price_ok = final_df[final_df['Price'].notna()].copy()
    df_price_ok['DateTime'] = df_price_ok.apply(lambda row: pd.Timestamp.combine(trade_date, row['Time']), axis=1)

    # Détection combinée des paires de Roll (priorité à 120 sec, sinon seuil étendu)
    roll_counter = 0
    for i in range(len(df_price_ok)):
        if df_price_ok.iloc[i]['Structure_ID'] != "":
            continue
        row_i = df_price_ok.iloc[i]

        # Vérification : si le ticker a exactement 7 caractères, c'est déjà un roll groupé
        if len(str(row_i['Ticker'])) == 7:
            roll_counter += 1
            roll_code = f"{trade_date.strftime('%Y%m%d')}-R-{roll_counter}-L0"
            idx_i = df_price_ok.index[i]
            final_df.loc[idx_i, 'Structure_ID'] = roll_code
            df_price_ok.loc[idx_i, 'Structure_ID'] = roll_code
            continue

        ticker_prefix_i = str(row_i['Ticker'])[:3]
        size_i = row_i['Size']
        price_i = row_i['Price']
        time_i = row_i['DateTime']

        candidate_j_120 = None
        candidate_j_extended = None

        for j in range(i+1, len(df_price_ok)):
            if df_price_ok.iloc[j]['Structure_ID'] != "":
                continue
            row_j = df_price_ok.iloc[j]
            if str(row_j['Ticker'])[:3] != ticker_prefix_i:
                continue
            if str(row_j['Ticker']) == str(row_i['Ticker']):
                continue
            if abs(size_i - row_j['Size']) > 0.05 * size_i:
                continue
            if abs(price_i - row_j['Price']) > 0.05 * price_i:
                continue

            time_diff = abs((row_j['DateTime'] - time_i).total_seconds())
            if time_diff <= 120:
                candidate_j_120 = j
                break
            elif time_diff <= 10000 and candidate_j_extended is None:
                candidate_j_extended = j

        if candidate_j_120 is not None:
            chosen_j = candidate_j_120
        elif candidate_j_extended is not None:
            chosen_j = candidate_j_extended
        else:
            continue

        roll_counter += 1
        roll_code_leg1 = f"{trade_date.strftime('%Y%m%d')}-R-{roll_counter}-L1"
        roll_code_leg2 = f"{trade_date.strftime('%Y%m%d')}-R-{roll_counter}-L2"
        idx_i = df_price_ok.index[i]
        idx_j = df_price_ok.index[chosen_j]
        final_df.loc[idx_i, 'Structure_ID'] = roll_code_leg1
        final_df.loc[idx_j, 'Structure_ID'] = roll_code_leg2
        df_price_ok.loc[idx_i, 'Structure_ID'] = roll_code_leg1
        df_price_ok.loc[idx_j, 'Structure_ID'] = roll_code_leg2

    # Attribution des labels "Screen" et "Outright"
    final_df.loc[final_df['Price'].isna(), 'Structure_ID'] = f"{trade_date.strftime('%Y%m%d')}-S"
    final_df.loc[final_df['Structure_ID'] == "", 'Structure_ID'] = f"{trade_date.strftime('%Y%m%d')}-O"

    # Création de la colonne "Structure"
    def extract_Structure(struct_code):
        if 'R' in struct_code:
            return 'Leg'
        elif 'S' in struct_code:
            return 'Screen'
        elif 'O' in struct_code:
            return 'Outright'
        else:
            return 'Autre'
    final_df['Structure'] = final_df['Structure_ID'].apply(extract_Structure)

    # Regroupement par catégorie et tri global
    order_mapping = {'Leg': 0, 'Screen': 1, 'Outright': 2, 'Autre': 3}
    final_df['sort_order'] = final_df['Structure'].map(order_mapping)
    final_df = final_df.sort_values(by='sort_order').drop(columns=['sort_order'])
    roll_mask = final_df['Structure_ID'].str.contains("-R-")
    screen_mask = final_df['Structure_ID'].str.contains("-S")
    outright_mask = final_df['Structure_ID'].str.contains("-O")
    df_roll = final_df[roll_mask].copy()
    df_screen = final_df[screen_mask].copy()
    df_outright = final_df[outright_mask].copy()

    # Tri personnalisé des Roll selon le ticker
    df_roll['roll_counter'] = df_roll['Structure_ID'].str.extract(r'-R-(\d+)-L').astype(int)
    df_roll['ticker_last_digit'] = df_roll['Ticker'].str[-1].astype(int)
    df_roll['ticker_penult'] = df_roll['Ticker'].str[-2]
    order_map_letters = {'H': 1, 'M': 2, 'U': 3, 'Z': 4}
    df_roll['ticker_penult_order'] = df_roll['ticker_penult'].map(order_map_letters).fillna(99)
    df_roll_sorted = df_roll.sort_values(by=['roll_counter', 'ticker_last_digit', 'ticker_penult_order'])
    df_screen_sorted = df_screen.sort_values(by='Time')
    df_outright_sorted = df_outright.sort_values(by='Time')
    final_sorted = pd.concat([df_roll_sorted, df_screen_sorted, df_outright_sorted], ignore_index=True)
    final_sorted = final_sorted.drop(columns=['roll_counter', 'ticker_last_digit', 'ticker_penult', 'ticker_penult_order'], errors='ignore')

    # Insertion des lignes résumé pour les Roll (Merge Roll) avec calcul "Level"
    summary_rows = []
    for r in df_roll_sorted['roll_counter'].unique():
        group = df_roll_sorted[df_roll_sorted['roll_counter'] == r]
        if len(group) >= 2:
            row1 = group.iloc[0]
            row2 = group.iloc[1]
            summary = {
                'Time': row1['Time'],
                'Level': (row2['Price'] / row1['Price'] - 1) * 100,
                'Ticker': row1['Ticker'] + row2['Ticker'][-2:],
                'Notional': (row1['Notional'] + row2['Notional']) / 2,
                'Size': row1['Size'],
                'Price': row1['Price'],
                'Volume': row1['Volume'],
                '1DChg': row1['1DChg'],
                'UndTkr': row1['UndTkr'],
                '1PtVal': row1['1PtVal'],
                'Exch': row1['Exch'],
                'UndCmpName': row1['UndCmpName'],
                'UndPrc': row1['UndPrc'],
                'Date': row1['Date'],
                'FutName': row1['FutName'] + row2['FutName'][-5:],
                'Structure_ID': f"{trade_date.strftime('%Y%m%d')}-R-{r}",
                'Structure': "Roll",
                'roll_counter': r,
                'order': 0
            }
            summary_rows.append(summary)
    df_summary = pd.DataFrame(summary_rows)
    legs_rows = []
    for _, row in df_roll_sorted.iterrows():
        row_copy = row.copy()
        row_copy['order'] = 1 if row_copy['Structure_ID'].endswith("-L1") else 2
        legs_rows.append(row_copy)
    df_legs = pd.DataFrame(legs_rows)
    df_roll_final = pd.concat([df_summary, df_legs], ignore_index=True)
    df_roll_final = df_roll_final.sort_values(by=['roll_counter', 'order'])
    df_roll_final = df_roll_final.drop(columns=['roll_counter', 'order', 'ticker_last_digit', 'ticker_penult', 'ticker_penult_order'], errors='ignore')
    mask_l0 = df_roll_final['Structure_ID'].str.contains("-L0", na=False)
    df_roll_final.loc[mask_l0, 'Structure_ID'] = df_roll_final.loc[mask_l0, 'Structure_ID'].str.replace("-L0", "", regex=False)
    df_roll_final.loc[mask_l0, 'Structure'] = "Roll"

    # Formatage de la date au format "MM/DD/YYYY"
    final_df['Date'] = pd.to_datetime(final_df['Date']).dt.strftime('%m/%d/%Y')

    # Assemblage final
    final_sorted = pd.concat([df_roll_final, df_screen_sorted, df_outright_sorted], ignore_index=True)

    # Insertion de la colonne "Closing1d" juste après "Price"
    price_idx = final_sorted.columns.get_loc("Price")
    final_sorted.insert(price_idx+1, "Closing1d", "")
    for i, row in final_sorted.iterrows():
        excel_row = i + 2  # Excel commence à la ligne 2 (après l'en-tête)
        if row["Structure"] in ["Roll", "Outright"]:
            final_sorted.at[i, "Closing1d"] = f'=BDH(J{excel_row}&" Index", "PX_CLOSE_1D",O{excel_row},O{excel_row})'
        else:
            final_sorted.at[i, "Closing1d"] = ""

    # Insertion de la formule dans la colonne "Level" pour les "Outright"
    # On insère ici une formule Excel qui calcule (Price/Closing1d - 1)
    for i, row in final_sorted.iterrows():
        if row["Structure"] == "Outright":
            excel_row = i + 2
            final_sorted.at[i, "Level"] = f'=(F{excel_row}/G{excel_row})'

    # Détection des Roll-Client : on vérifie uniquement les lignes de type Leg pour déterminer si les 2 legs ont le même Price
    roll_groups = final_sorted[final_sorted['Structure'].isin(['Roll', 'Leg'])].groupby(
        final_sorted['Structure_ID'].str.extract(r'(\d{8}-R-\d+)')[0]
    )
    for roll_id, group in roll_groups:
        # Récupérer uniquement les lignes de type Leg
        group_legs = group[group['Structure'] == 'Leg']
        if len(group_legs) == 2 and group_legs['Price'].nunique() == 1:
            # Mise à jour uniquement de la ligne de synthèse (summary) ayant Structure "Roll"
            summary_indices = group[group['Structure'] == 'Roll'].index
            final_sorted.loc[summary_indices, 'Structure'] = 'Roll-Client'

    return final_sorted

def postprocess_excel(final_sorted):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        final_sorted.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    closing1d_col_idx = final_sorted.columns.get_loc("Closing1d") + 1  # openpyxl est 1-indexé
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=closing1d_col_idx)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            ws.cell(row=row_num, column=closing1d_col_idx).value = f'=BDH(J{row_num}&" Index", "PX_CLOSE_1D",O{row_num},O{row_num})'
    level_col_idx = final_sorted.columns.get_loc("Level") + 1 if "Level" in final_sorted.columns else None
    if level_col_idx:
        for row_num in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=level_col_idx)
            cell.number_format = "0.000"
    download_buffer = io.BytesIO()
    wb.save(download_buffer)
    download_buffer.seek(0)
    return download_buffer

def save_processed_data(new_data, filename="data/processed/processed_data.csv"):
    """
    Pour une date donnée, si des données existent déjà dans le fichier de sauvegarde,
    on supprime les lignes correspondantes avant d'ajouter les nouvelles.
    """
    if not os.path.exists("data/processed"):
        os.makedirs("data/processed")
    if os.path.exists(filename):
        try:
            old_data = pd.read_csv(filename)
        except Exception as e:
            old_data = pd.DataFrame()
    else:
        old_data = pd.DataFrame()
    # Récupérer la date traitée (toutes les lignes de new_data concernent la même date)
    processed_date = new_data['Date'].iloc[0]
    if 'Date' in old_data.columns:
        old_data = old_data[old_data['Date'] != processed_date]
    combined = pd.concat([old_data, new_data], ignore_index=True)
    combined.to_csv(filename, index=False)
    return combined

# ------------------------
# Fonctions pour le calendrier
# ------------------------

def generate_month_calendar(year, month, processed_dates_set):
    """
    Génère un calendrier HTML pour un mois donné.
    Les dates présentes dans processed_dates_set sont surlignées en rouge avec du texte en gras.
    """
    month_days = calendar.monthcalendar(year, month)
    html = f'<table border="1" style="border-collapse: collapse; text-align: center; margin: 10px;">'
    html += f'<tr><th colspan="7" style="background-color: #ddd;">{calendar.month_name[month]} {year}</th></tr>'
    # En-tête des jours (abréviations en français)
    jours = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    html += "<tr>" + "".join(f"<th style='padding: 2px 6px; background-color: #f0f0f0;'>{j}</th>" for j in jours) + "</tr>"
    for week in month_days:
        html += "<tr>"
        for day in week:
            if day == 0:
                html += "<td>&nbsp;</td>"
            else:
                d = date(year, month, day)
                if d in processed_dates_set:
                    html += f'<td style="padding: 4px; background-color: red; color: white; font-weight: bold;">{day}</td>'
                else:
                    html += f'<td style="padding: 4px;">{day}</td>'
        html += "</tr>"
    html += "</table>"
    return html

def generate_calendar_html(processed_dates_set):
    """
    Génère un bloc HTML affichant les calendriers des trois derniers mois.
    """
    html_parts = []
    today = date.today()
    months = []
    # Calculer les 3 derniers mois (incluant le mois courant)
    for i in range(3):
        m = today.month - i
        y = today.year
        if m <= 0:
            m += 12
            y -= 1
        months.append((y, m))
    months = sorted(months)  # Afficher dans l'ordre chronologique
    for (y, m) in months:
        html_parts.append(generate_month_calendar(y, m, processed_dates_set))
    return '<div style="display: flex; flex-wrap: wrap;">' + "".join(html_parts) + '</div>'

# ------------------------
# Fonction principale
# ------------------------

def main():
    st.title("Application de traitement des fichiers Excel")

    uploaded_files = st.file_uploader("Sélectionnez un ou plusieurs fichiers Excel", type=["xlsx"], accept_multiple_files=True)
    trade_date = st.date_input("Sélectionnez la date")

    if uploaded_files and trade_date:
        if st.button("Traiter les fichiers"):
            with st.spinner("Traitement en cours..."):
                processed_df = process_files(uploaded_files, pd.to_datetime(trade_date))
                if processed_df is not None:
                    # Sauvegarde dans data/processed (remplacement des données pour la date donnée)
                    save_processed_data(processed_df)
                    download_buffer = postprocess_excel(processed_df)
                    st.success("Traitement terminé!")
                    st.download_button(
                        label="Télécharger le fichier traité",
                        data=download_buffer,
                        file_name="fichier_traite.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.dataframe(processed_df)

    # ------------------------
    # Affichage du calendrier en bas de page
    # ------------------------
    st.markdown("---")
    st.markdown("### Calendrier des trois derniers mois")
    processed_dates_set = set()
    global_file = "data/processed/processed_data.csv"
    if os.path.exists(global_file):
        try:
            df_global = pd.read_csv(global_file)
            # On suppose que la colonne 'Date' est au format MM/DD/YYYY
            for d_str in df_global['Date'].unique():
                try:
                    d_str = d_str.strip()  # nettoyage des espaces éventuels
                    d_obj = datetime.strptime(d_str, "%m/%d/%Y").date()
                    processed_dates_set.add(d_obj)
                except Exception as e:
                    pass
        except Exception as e:
            st.error("Erreur lors du chargement du fichier global.")
    else:
        st.info("Aucune donnée traitée pour l'instant.")

    calendar_html = generate_calendar_html(processed_dates_set)
    st.markdown(calendar_html, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
