import io
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def load_dataframes(file_sources, on_error=None):
    dataframes = []
    for src in file_sources:
        name = getattr(src, 'name', str(src))
        try:
            dataframes.append(pd.read_excel(src, engine='openpyxl'))
        except Exception as e:
            if on_error:
                on_error(f"Erreur lors du chargement de {name}: {e}")
    return dataframes


def process_data(dataframes, trade_date):
    """
    Traite une liste de DataFrames Bloomberg et retourne (DataFrame, erreur).
    trade_date : datetime.date ou pd.Timestamp
    """
    final_df = pd.concat(dataframes, ignore_index=True)

    if 'Time' not in final_df.columns:
        return None, "La colonne 'Time' est introuvable dans les fichiers."

    final_df['Time'] = pd.to_datetime(
        final_df['Time'], format='%H:%M:%S', errors='coerce'
    ).dt.time

    cutoff = pd.Timestamp('08:00:00').time()
    final_df['_sort'] = (final_df['Time'] < cutoff).astype(int)
    final_df = (
        final_df.sort_values(['_sort', 'Time'])
                .drop(columns=['_sort'])
                .reset_index(drop=True)
                .iloc[::-1]
                .reset_index(drop=True)
    )

    trade_date = pd.Timestamp(trade_date).date()
    date_str = pd.Timestamp(trade_date).strftime('%Y%m%d')

    final_df['Date'] = trade_date
    final_df['Structure_ID'] = ''
    final_df['Price'] = pd.to_numeric(final_df['Price'], errors='coerce')
    final_df['Size'] = pd.to_numeric(final_df['Size'], errors='coerce')

    df_price_ok = final_df[final_df['Price'].notna()].copy()
    df_price_ok['DateTime'] = pd.to_datetime(
        pd.Timestamp(trade_date).strftime('%Y-%m-%d') + ' ' + df_price_ok['Time'].astype(str)
    )

    # --- Détection des Rolls (O(n²) stateful, pas vectorisable) ---
    # Utilise des arrays numpy pour minimiser le coût d'accès pandas
    ticker_arr = df_price_ok['Ticker'].astype(str).values
    size_arr = df_price_ok['Size'].values
    price_arr = df_price_ok['Price'].values
    datetime_arr = df_price_ok['DateTime'].values
    struct_arr = df_price_ok['Structure_ID'].values.copy()

    roll_counter = 0
    for i in range(len(df_price_ok)):
        if struct_arr[i] != '':
            continue
        ticker_i = ticker_arr[i]
        if len(ticker_i) == 7:
            roll_counter += 1
            struct_arr[i] = f'{date_str}-R-{roll_counter}-L0'
            continue

        prefix_i = ticker_i[:3]
        size_i, price_i, time_i = size_arr[i], price_arr[i], datetime_arr[i]
        cand_120 = cand_ext = None

        for j in range(i + 1, len(df_price_ok)):
            if struct_arr[j] != '':
                continue
            tj = ticker_arr[j]
            if tj[:3] != prefix_i or tj == ticker_i:
                continue
            if abs(size_i - size_arr[j]) > 0.05 * size_i:
                continue
            if abs(price_i - price_arr[j]) > 0.05 * price_i:
                continue
            diff = abs(int((datetime_arr[j] - time_i) / np.timedelta64(1, 's')))
            if diff <= 120:
                cand_120 = j
                break
            elif diff <= 10000 and cand_ext is None:
                cand_ext = j

        chosen = cand_120 if cand_120 is not None else cand_ext
        if chosen is None:
            continue
        roll_counter += 1
        struct_arr[i] = f'{date_str}-R-{roll_counter}-L1'
        struct_arr[chosen] = f'{date_str}-R-{roll_counter}-L2'

    df_price_ok['Structure_ID'] = struct_arr
    final_df.loc[df_price_ok.index, 'Structure_ID'] = struct_arr

    # --- Labels Screen / Outright ---
    final_df.loc[final_df['Price'].isna(), 'Structure_ID'] = f'{date_str}-S'
    final_df.loc[final_df['Structure_ID'] == '', 'Structure_ID'] = f'{date_str}-O'

    # --- Colonne Structure (vectorisé) ---
    final_df['Structure'] = np.select(
        [
            final_df['Structure_ID'].str.contains('-R-'),
            final_df['Structure_ID'].str.contains('-S'),
            final_df['Structure_ID'].str.contains('-O'),
        ],
        ['Leg', 'Roll Screen', 'Outright'],
        default='Autre',
    )

    df_roll = final_df[final_df['Structure_ID'].str.contains('-R-')].copy()
    df_screen = final_df[final_df['Structure_ID'].str.contains('-S')].copy()
    df_outright = final_df[final_df['Structure_ID'].str.contains('-O')].copy()

    # --- Tri des Rolls ---
    df_roll['roll_counter'] = pd.to_numeric(
        df_roll['Structure_ID'].str.extract(r'-R-(\d+)-L', expand=False), errors='coerce'
    ).fillna(0).astype(int)
    df_roll['ticker_last_digit'] = pd.to_numeric(
        df_roll['Ticker'].astype(str).str[-1], errors='coerce'
    ).fillna(0).astype(int)
    df_roll['ticker_penult'] = df_roll['Ticker'].str[-2]
    df_roll['ticker_penult_order'] = df_roll['ticker_penult'].map(
        {'H': 1, 'M': 2, 'U': 3, 'Z': 4}
    ).fillna(99)
    df_roll_sorted = df_roll.sort_values(
        ['roll_counter', 'ticker_last_digit', 'ticker_penult_order']
    )

    # --- Lignes résumé Roll ---
    summary_rows = []
    for r, group in df_roll_sorted.groupby('roll_counter'):
        if len(group) >= 2:
            r1, r2 = group.iloc[0], group.iloc[1]
            summary_rows.append({
                'Time': r1['Time'],
                'Level': (r2['Price'] / r1['Price'] - 1) * 100,
                'Ticker': r1['Ticker'] + r2['Ticker'][-2:],
                'Notional': (r1['Notional'] + r2['Notional']) / 2,
                'Size': r1['Size'], 'Price': r1['Price'],
                'Volume': r1['Volume'], '1DChg': r1['1DChg'],
                'UndTkr': r1['UndTkr'], '1PtVal': r1['1PtVal'],
                'Exch': r1['Exch'], 'UndCmpName': r1['UndCmpName'],
                'UndPrc': r1['UndPrc'], 'Date': r1['Date'],
                'FutName': r1['FutName'] + r2['FutName'][-5:],
                'Structure_ID': f'{date_str}-R-{r}',
                'Structure': 'Roll', 'roll_counter': r, 'order': 0,
            })

    df_legs = df_roll_sorted.copy()
    df_legs['order'] = df_legs['Structure_ID'].str.endswith('-L1').map(
        {True: 1, False: 2}
    )

    sort_cols = ['roll_counter', 'order', 'ticker_last_digit', 'ticker_penult', 'ticker_penult_order']
    df_roll_final = (
        pd.concat([pd.DataFrame(summary_rows), df_legs], ignore_index=True)
          .sort_values(['roll_counter', 'order'])
          .drop(columns=sort_cols, errors='ignore')
    )

    mask_l0 = df_roll_final['Structure_ID'].str.contains('-L0', na=False)
    df_roll_final.loc[mask_l0, 'Structure_ID'] = (
        df_roll_final.loc[mask_l0, 'Structure_ID'].str.replace('-L0', '', regex=False)
    )
    df_roll_final.loc[mask_l0, 'Structure'] = 'Roll'

    final_df['Date'] = pd.to_datetime(final_df['Date']).dt.strftime('%m/%d/%Y')

    final_sorted = pd.concat([
        df_roll_final,
        df_screen.sort_values('Time'),
        df_outright.sort_values('Time'),
    ], ignore_index=True)

    if 'Level' not in final_sorted.columns:
        final_sorted['Level'] = ''

    # --- Colonne Closing1d (vectorisé) ---
    price_idx = final_sorted.columns.get_loc('Price')
    final_sorted.insert(price_idx + 1, 'Closing1d', '')
    excel_rows = pd.Series(
        range(2, len(final_sorted) + 2), index=final_sorted.index
    )
    mask_closing = final_sorted['Structure'].isin(['Roll', 'Outright'])
    final_sorted.loc[mask_closing, 'Closing1d'] = excel_rows[mask_closing].apply(
        lambda r: f'=BDH(L{r}&" Index", "PX_CLOSE_1D",R{r},R{r})'
    )

    # --- Formule Level pour Outright (vectorisé) ---
    mask_out = final_sorted['Structure'] == 'Outright'
    final_sorted.loc[mask_out, 'Level'] = excel_rows[mask_out].apply(
        lambda r: f'=(F{r}/G{r})'
    )

    # --- Détection Roll Client (price identique OU notional identique) ---
    final_sorted = _detect_roll_clients(final_sorted)
    return _reorder_columns(final_sorted), None


def _detect_roll_clients(df):
    df = df.copy()
    df['_group'] = df['Structure_ID'].str.extract(r'(\d{8}-R-\d+)')
    for _, group in df[df['Structure'].isin(['Roll', 'Leg'])].groupby('_group'):
        legs = group[group['Structure'] == 'Leg']
        roll = group[group['Structure'] == 'Roll']
        if len(legs) == 2 and not roll.empty:
            if legs['Price'].nunique() == 1 or legs['Notional'].nunique() == 1:
                df.loc[roll.index, 'Structure'] = 'Roll Client'
    return df.drop(columns=['_group'])


def _reorder_columns(df):
    desired = [
        'Time', 'Level', 'Ticker', 'Notional', 'Size', 'Price', 'Closing1d',
        'Structure_ID', 'Structure', 'Volume', '1DChg', 'UndTkr', '1PtVal',
        'Exch', 'FutName', 'UndCmpName', 'UndPrc', 'Date',
    ]
    others = [c for c in df.columns if c not in desired]
    return df[desired + others]


def build_excel(final_sorted):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        final_sorted.to_excel(writer, index=False, sheet_name='Sheet1')
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active
    closing_col = final_sorted.columns.get_loc('Closing1d') + 1
    level_col = (
        final_sorted.columns.get_loc('Level') + 1
        if 'Level' in final_sorted.columns else None
    )

    for row_num in range(4, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=closing_col)
        if isinstance(cell.value, str) and cell.value.startswith('='):
            cell.value = f'=BDH(L{row_num}&" Index", "PX_CLOSE_1D",R{row_num},R{row_num})'

    if level_col:
        for row_num in range(2, ws.max_row + 1):
            ws.cell(row=row_num, column=level_col).number_format = '0.000'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
